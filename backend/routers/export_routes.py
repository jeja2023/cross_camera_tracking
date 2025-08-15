from fastapi import APIRouter, Depends, UploadFile, File, BackgroundTasks, Form, HTTPException, status, Query, Body
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
import io
import os
import datetime
from typing import List, Optional
import logging
from urllib.parse import quote
import aiofiles
import shutil
import tempfile

from ..database_conn import get_db
from .. import crud
from .. import schemas
from .. import auth
from ..ml_services import ml_logic # Import ml_logic
from ..config import settings # Import settings

router = APIRouter(
    tags=["导出"],
    responses={404: {"description": "未找到"}},
)

logger = logging.getLogger(__name__)

@router.get("/video_results/{video_id}", summary="导出特定视频的人物特征图库结果到Excel")
async def export_video_results(
    video_id: int,
    db: Session = Depends(get_db),
    current_user: schemas.User = Depends(auth.get_current_active_user)
):
    logger.info(f"用户 {current_user.username} 正在导出视频 (ID: {video_id}) 的解析结果到Excel。")
    video = crud.get_video(db, video_id=video_id)
    if not video:
        logger.warning(f"用户 {current_user.username} 尝试导出不存在的视频 (ID: {video_id}) 的结果。")
        raise HTTPException(status_code=404, detail="视频未找到")
    
    # 权限检查
    if video.owner_id != current_user.id and current_user.role != "admin":
        logger.warning(f"用户 {current_user.username} (角色: {current_user.role}) 未经授权尝试导出视频 (ID: {video_id}) 的结果。")
        raise HTTPException(status_code=403, detail="无足够权限")
    
    persons = crud.get_persons_by_video_id(db, video_id=video_id)
    logger.info(f"视频 (ID: {video_id}) 找到 {len(persons)} 个人物数据用于导出。")
    
    # 限制最多导出图片数量，从配置中获取
    persons_to_export = persons[:settings.EXCEL_EXPORT_MAX_IMAGES]
    logger.info(f"实际导出前 {len(persons_to_export)} 张图片。")
    
    # 创建新的Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = f"视频_{video_id}_结果"

    # 设置表头
    headers = ["序号", "视频名称", "视频UUID", "图片ID", "图片缩小图"]
    ws.append(headers)
    # 设置E列的宽度以容纳图片
    ws.column_dimensions['E'].width = 15 # 假设图片宽度为100像素，约等于15个字符宽度
    # 设置C列的宽度以容纳UUID
    ws.column_dimensions['C'].width = 38 # 视频UUID可能较长，给予足够宽度
    logger.info("Excel 表头已设置。")

    # 写入数据
    for idx, p in enumerate(persons_to_export):
        # 序号从1开始
        # 确保 video 对象是可用的，并且包含 filename 和 uuid 属性
        video_filename = video.filename if video and hasattr(video, 'filename') else "N/A"
        video_uuid = video.uuid if video and hasattr(video, 'uuid') else "N/A"
        row_data = [idx + 1, video_filename, video_uuid, p.uuid]
        ws.append(row_data)
        logger.info(f"正在处理图片: {p.crop_image_path} (UUID: {p.uuid}, 视频名称: {video_filename}, 视频UUID: {video_uuid})")

        # 添加图片
        # 检查 p.crop_image_path 是否存在
        if not p.crop_image_path:
            logger.warning(f"人物 {p.uuid} 没有 crop_image_path，跳过添加图片。")
            ws.append(["", "", "", "", "无缩小图"]) # 添加占位符
            continue

        # 构建图片绝对路径
        relative_path_from_crops_root = p.crop_image_path
        if relative_path_from_crops_root.startswith("database/crops/"):
            relative_path_from_crops_root = relative_path_from_crops_root[len("database/crops/"):]
        elif relative_path_from_crops_root.startswith("database\\crops\\"):
            relative_path_from_crops_root = relative_path_from_crops_root[len("database\\crops\\"):]

        image_path = os.path.join(settings.DATABASE_CROPS_DIR, relative_path_from_crops_root)
        logger.info(f"尝试为人物 {p.uuid} 加载图片: {image_path}")
        
        # 检查图片文件是否存在
        if not os.path.exists(image_path):
            logger.warning(f"图片文件未找到: {image_path}. 跳过此图片。")
            ws.append(["", "", "", "", "图片文件不存在"]) # 添加占位符
            continue

        try:
            pil_img = PILImage.open(image_path)
            logger.info(f"成功打开图片: {image_path}, 模式: {pil_img.mode}, 尺寸: {pil_img.size}")
            # 缩小图片到合适大小，从配置中获取
            pil_img.thumbnail((settings.EXCEL_EXPORT_IMAGE_SIZE_PX, settings.EXCEL_EXPORT_IMAGE_SIZE_PX)) 
            logger.info(f"图片 {image_path} 已缩小到 {pil_img.size[0]}x{pil_img.size[1]}。")
            
            img_byte_arr = io.BytesIO()
            # 确保图片为RGB模式并保存为PNG格式，以提高兼容性
            pil_img.convert("RGB").save(img_byte_arr, format='PNG')
            img_byte_arr.seek(0)
            logger.info(f"图片 {image_path} 已保存到内存，大小: {len(img_byte_arr.getvalue())} 字节。")

            excel_img = ExcelImage(img_byte_arr)
            excel_img.width = pil_img.width
            excel_img.height = pil_img.height
            
            # 将图片插入到第五列，即 E 列
            img_cell = f"E{ws.max_row}"
            ws.add_image(excel_img, img_cell)
            # 设置行高以容纳图片，从配置中获取
            ws.row_dimensions[ws.max_row].height = settings.EXCEL_EXPORT_ROW_HEIGHT_PT # 约等于100像素
            logger.info(f"图片 {image_path} 已添加到Excel的 {img_cell} 单元格。")

        except Exception as e:
            logger.error(f"处理图片 {image_path} 失败: {e}", exc_info=True)
            ws.append(["", "", "", "", f"处理失败: {e}"]) # 添加占位符

    # 创建临时文件来保存Excel工作簿
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    logger.info(f"视频 (ID: {video_id}) 的Excel报告已生成。")
    
    # 设置响应头
    file_name = f"视频_{video_id}_结果_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    logger.info(f"生成的视频导出文件名: {file_name}")
    
    # 生成通用ASCII回退文件名
    generic_filename = "video_export_report.xlsx"
    # 对包含中文和时间戳的完整文件名进行URL编码，用于filename*参数
    encoded_full_filename = quote(file_name)

    headers = {
        'Content-Disposition': f"""attachment; filename="{generic_filename}"; filename*=UTF-8''{encoded_full_filename}""",
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    }
    logger.debug(f"导出视频结果 - 原始文件名: {file_name}, 通用文件名: {generic_filename}, 编码文件名 (for filename*): {encoded_full_filename}, 响应头: {headers}")
    return StreamingResponse(io.BytesIO(excel_file.getvalue()), headers=headers, media_type="application/octet-stream")

@router.post("/query_results", summary="导出以图搜人查询结果到Excel")
async def export_query_results(
    db: Session = Depends(get_db),
    threshold: float = Form(0.7),
    query_image: UploadFile = File(...),
    video_uuid: Optional[str] = Form(None),
    stream_uuid: Optional[str] = Form(None),
    current_user: schemas.User = Depends(auth.get_current_active_user),
    background_tasks: BackgroundTasks = None
):
    logger.info(f"用户 {current_user.username} 正在导出以图搜人查询结果。阈值: {threshold}, 视频UUID: {video_uuid if video_uuid else '全局'}, 视频流UUID: {stream_uuid if stream_uuid else '全局'}")

    # 保存查询图片到临时文件
    query_image_path = os.path.join(settings.UPLOAD_DIR, query_image.filename)
    async with aiofiles.open(query_image_path, "wb") as out_file:
        content = await query_image.read()
        await out_file.write(content)
    
    # 将临时图片文件添加到后台任务，以便在请求完成后删除
    background_tasks.add_task(os.remove, query_image_path)

    # 调试日志：打印即将传递给ml_logic的query_image_path
    logger.debug(f"export_query_results: 准备传递给 find_similar_people 的查询图片路径: {query_image_path}")

    # 执行以图搜人查询，获取所有结果
    # 注意：这里调用 ml_logic.find_similar_people 将不带 skip 和 limit，以获取所有匹配项
    search_results = ml_logic.find_similar_people(db, query_image_path, threshold, video_uuid=video_uuid, stream_uuid=stream_uuid, current_user=current_user)
    total_results = len(search_results["items"])
    logger.info(f"以图搜人查询完成，共找到 {total_results} 个结果用于导出。")

    if total_results == 0:
        logger.info(f"以图搜人查询无结果，不生成Excel文件。")
        raise HTTPException(status_code=204, detail="没有搜索结果可导出") # 204 No Content
    
    # 创建新的Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "图片搜索结果"

    # 设置表头
    headers = ["序号", "人物ID", "相似度", "来源类型", "来源名称", "来源UUID", "裁剪图片", "完整帧图片"]
    ws.append(headers)
    # 设置对应列的宽度以容纳图片和文本
    ws.column_dimensions['F'].width = 38 # 来源UUID可能较长，给予足够宽度
    ws.column_dimensions['G'].width = 15 # 裁剪图片列的宽度
    ws.column_dimensions['H'].width = 15 # 完整帧图片列的宽度
    logger.info("Excel 表头已设置。")

    # 扁平化搜索结果，处理每个查询人物的匹配结果
    all_matched_persons = []
    for grouped_result in search_results["items"]:
        all_matched_persons.extend(grouped_result["results"])
    
    # 可以选择根据相似度再次排序所有结果（可选）
    all_matched_persons.sort(key=lambda x: x.get('score', 0.0), reverse=True)

    # 写入数据
    for idx, result in enumerate(all_matched_persons):
        logger.debug(f"export_query_results: Processing result item: {result}")
        source_type = "未知"
        source_name = "N/A"
        source_uuid = "N/A"

        if result.get('stream_uuid'):
            source_type = "视频流"
            source_name = result.get('stream_name') or "N/A"
            source_uuid = result.get('stream_uuid')
        elif result.get('video_uuid'):
            source_type = "上传视频"
            source_name = result.get('video_filename') or "N/A"
            source_uuid = result.get('video_uuid')
        elif result.get('upload_image_uuid'): # 新增对上传图片的来源支持
            source_type = "上传图片"
            source_name = result.get('upload_image_filename') or "N/A"
            source_uuid = result.get('upload_image_uuid')

        row_data = [
            idx + 1,
            result.get('uuid', 'N/A'),
            f"{result.get('score', 0.0):.2f}%", # 格式化相似度为百分比，移除多余的 * 100
            source_type,
            source_name,
            source_uuid,
            '', # 裁剪图片占位符
            ''  # 完整帧图片占位符
        ]
        ws.append(row_data)
        logger.info(f"正在处理导出结果中的人物: {result.get('uuid', 'N/A')}，相似度: {result.get('score', 0.0):.2f}%")

        # 获取当前行的索引，图片将插入到这一行
        current_row = ws.max_row

        # 添加裁剪图片 (第7列，索引6)
        if result.get('crop_image_path') and os.path.exists(os.path.join(settings.BASE_DIR, "backend", result['crop_image_path'])):
            try:
                crop_image_full_path = os.path.join(settings.BASE_DIR, "backend", result['crop_image_path'])
                pil_img_crop = PILImage.open(crop_image_full_path)
                logger.info(f"成功打开裁剪图片: {crop_image_full_path}, 模式: {pil_img_crop.mode}, 尺寸: {pil_img_crop.size}")
                pil_img_crop.thumbnail((settings.EXCEL_EXPORT_IMAGE_SIZE_PX, settings.EXCEL_EXPORT_IMAGE_SIZE_PX)) 
                logger.info(f"裁剪图片 {crop_image_full_path} 已缩小到 {pil_img_crop.size[0]}x{pil_img_crop.size[1]}。")
                
                img_byte_arr_crop = io.BytesIO()
                pil_img_crop.convert("RGB").save(img_byte_arr_crop, format='PNG')
                img_byte_arr_crop.seek(0)
                logger.info(f"裁剪图片 {crop_image_full_path} 已保存到内存，大小: {len(img_byte_arr_crop.getvalue())} 字节。")

                excel_img_crop = ExcelImage(img_byte_arr_crop)
                excel_img_crop.width = pil_img_crop.width
                excel_img_crop.height = pil_img_crop.height
                
                # 将裁剪图片插入到G列
                img_cell_crop = f"G{current_row}"
                ws.add_image(excel_img_crop, img_cell_crop)
                logger.info(f"裁剪图片 {crop_image_full_path} 已添加到Excel的 {img_cell_crop} 单元格。")

            except Exception as e:
                logger.error(f"处理裁剪图片 {result.get('crop_image_path', 'N/A')} 失败: {e}", exc_info=True)
                ws.cell(row=current_row, column=7, value=f"处理失败: {e}") # 直接写入到G列
        else:
            logger.warning(f"人物 {result.get('uuid', 'N/A')} 没有 crop_image_path 或文件不存在: {result.get('crop_image_path', 'N/A')}. 跳过添加裁剪图片。")
            ws.cell(row=current_row, column=7, value="无缩小图") # 直接写入到G列
        
        # 添加完整帧图片 (第8列，索引7)
        if result.get('full_frame_image_path') and os.path.exists(os.path.join(settings.BASE_DIR, "backend", result['full_frame_image_path'])):
            try:
                full_frame_image_full_path = os.path.join(settings.BASE_DIR, "backend", result['full_frame_image_path'])
                pil_img_full = PILImage.open(full_frame_image_full_path)
                logger.info(f"成功打开完整帧图片: {full_frame_image_full_path}, 模式: {pil_img_full.mode}, 尺寸: {pil_img_full.size}")
                pil_img_full.thumbnail((settings.EXCEL_EXPORT_IMAGE_SIZE_PX, settings.EXCEL_EXPORT_IMAGE_SIZE_PX))
                logger.info(f"完整帧图片 {full_frame_image_full_path} 已缩小到 {pil_img_full.size[0]}x{pil_img_full.size[1]}。")
                
                img_byte_arr_full = io.BytesIO()
                pil_img_full.convert("RGB").save(img_byte_arr_full, format='PNG')
                img_byte_arr_full.seek(0)
                logger.info(f"完整帧图片 {full_frame_image_full_path} 已保存到内存，大小: {len(img_byte_arr_full.getvalue())} 字节。")

                excel_img_full = ExcelImage(img_byte_arr_full)
                excel_img_full.width = pil_img_full.width
                excel_img_full.height = pil_img_full.height
                
                # 将完整帧图片插入到H列
                img_cell_full = f"H{current_row}"
                ws.add_image(excel_img_full, img_cell_full)
                logger.info(f"完整帧图片 {full_frame_image_full_path} 已添加到Excel的 {img_cell_full} 单元格。")

            except Exception as e:
                logger.error(f"处理完整帧图片 {result.get('full_frame_image_path', 'N/A')} 失败: {e}", exc_info=True)
                ws.cell(row=current_row, column=8, value=f"处理失败: {e}") # 直接写入到H列
        else:
            logger.warning(f"人物 {result.get('uuid', 'N/A')} 没有 full_frame_image_path 或文件不存在: {result.get('full_frame_image_path', 'N/A')}. 跳过添加完整帧图片。")
            ws.cell(row=current_row, column=8, value="无完整帧图") # 直接写入到H列
        
        # 设置行高以容纳图片
        ws.row_dimensions[current_row].height = settings.EXCEL_EXPORT_ROW_HEIGHT_PT # 约等于100像素

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    logger.info(f"以图搜人Excel报告已生成。")
    
    file_name = f"以图搜人结果_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    logger.info(f"生成的以图搜人导出文件名: {file_name}")
    
    generic_filename = "image_search_export_report.xlsx"
    encoded_full_filename = quote(file_name)

    headers = {
        'Content-Disposition': f"attachment; filename=\"{generic_filename}\"; filename*=UTF-8''{encoded_full_filename}\"",
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    }
    logger.debug(f"导出以图搜人结果 - 原始文件名: {file_name}, 通用文件名: {generic_filename}, 编码文件名 (for filename*): {encoded_full_filename}, 响应头: {headers}")
    return StreamingResponse(io.BytesIO(excel_file.getvalue()), headers=headers, media_type="application/octet-stream") 

@router.get("/all_persons", summary="导出全部人物特征图库到Excel (管理员权限，支持筛选)")
async def export_all_persons(
    db: Session = Depends(get_db),
    is_verified: Optional[bool] = Query(None, description="按审核状态筛选 (True: 已审核, False: 未审核)"),
    marked_for_retrain: Optional[bool] = Query(None, description="按再训练标记筛选 (True: 待再训练, False: 不待再训练)"),
    query: Optional[str] = Query(None, description="根据人物UUID、视频UUID、视频流UUID、图片UUID或名称进行模糊搜索"),
    has_id_card: Optional[bool] = Query(None, description="筛选是否有身份证号/ID数据 (True: 有, False: 无)"), # 新增参数
    current_user: schemas.User = Depends(auth.get_current_active_user)
):
    if current_user.role != "admin" and current_user.role != "advanced":
        logger.warning(f"用户 {current_user.username} (角色: {current_user.role}) 未经授权尝试导出全部人物特征图库。")
        raise HTTPException(status_code=403, detail="Not enough permissions")

    # 获取所有符合筛选条件的人物数据 (不分页)
    persons = crud.get_persons(db, skip=0, limit=None, is_verified=is_verified, marked_for_retrain=marked_for_retrain, query=query, has_id_card=has_id_card)

    if not persons:
        logger.info("没有符合筛选条件的人物特征可导出。")
        raise HTTPException(status_code=204, detail="没有符合筛选条件的人物特征可导出")
    
    # 创建新的Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "全部人物特征"

    # 设置表头
    headers = ["序号", "人物UUID", "裁剪图像", "完整帧图像", "来源类型", "来源名称", "来源UUID", "入库时间"]
    ws.append(headers)

    # 设置列宽
    ws.column_dimensions['B'].width = 38 # 人物UUID
    ws.column_dimensions['C'].width = 15 # 裁剪图像
    ws.column_dimensions['D'].width = 15 # 完整帧图像
    ws.column_dimensions['F'].width = 30 # 来源名称
    ws.column_dimensions['G'].width = 38 # 来源UUID
    ws.column_dimensions['H'].width = 25 # 入库时间
    # 移除置信度、是否审核、是否待再训练的列宽设置
    # ws.column_dimensions['I'].width = 12 # 置信度
    # ws.column_dimensions['J'].width = 10 # 是否审核
    # ws.column_dimensions['K'].width = 15 # 是否待再训练

    logger.info("Excel 表头和列宽已设置。")

    # 写入数据
    for idx, person in enumerate(persons):
        source_type = "未知"
        source_name = "N/A"
        source_uuid = "N/A"

        if person.stream_uuid:
            source_type = "视频流"
            source_name = person.stream_name or "N/A"
            source_uuid = person.stream_uuid
        elif person.video_uuid:
            source_type = "上传视频"
            source_name = person.video_name or "N/A"
            source_uuid = person.video_uuid
        elif person.upload_image_uuid:
            source_type = "上传图片"
            source_name = person.upload_image_filename or "N/A"
            source_uuid = person.upload_image_uuid
        
        display_time = person.created_at.strftime('%Y-%m-%d %H:%M:%S') if person.created_at else 'N/A'
        # confidence_score = f"{person.confidence_score * 100:.2f}%" if person.confidence_score is not None else 'N/A' # 移除

        row_data = [
            idx + 1,
            person.uuid,
            '', # 裁剪图像占位符
            '', # 完整帧图像占位符
            source_type,
            source_name,
            source_uuid,
            display_time
            # 移除 confidence_score, is_verified, marked_for_retrain
            # confidence_score,
            # "是" if person.is_verified else "否",
            # "是" if person.marked_for_retrain else "否"
        ]
        ws.append(row_data)

        # 添加裁剪图像
        if person.crop_image_path:
            # 移除 person.crop_image_path 中重复的 'database/crops/' 前缀
            relative_path_from_crops_root = person.crop_image_path
            if relative_path_from_crops_root.startswith("database/crops/"):
                relative_path_from_crops_root = relative_path_from_crops_root[len("database/crops/"):]
            elif relative_path_from_crops_root.startswith("database\\crops\\"):
                relative_path_from_crops_root = relative_path_from_crops_root[len("database\\crops\\"):]
            
            crop_image_full_path = os.path.join(settings.DATABASE_CROPS_DIR, relative_path_from_crops_root)
            logger.info(f"尝试为人物 {person.uuid} 加载裁剪图片: {crop_image_full_path}")

            if os.path.exists(crop_image_full_path):
                try:
                    pil_img_crop = PILImage.open(crop_image_full_path)
                    pil_img_crop.thumbnail((settings.EXCEL_EXPORT_IMAGE_SIZE_PX, settings.EXCEL_EXPORT_IMAGE_SIZE_PX))
                    img_byte_arr_crop = io.BytesIO()
                    pil_img_crop.convert("RGB").save(img_byte_arr_crop, format='PNG')
                    img_byte_arr_crop.seek(0)
                    excel_img_crop = ExcelImage(img_byte_arr_crop)
                    excel_img_crop.width = pil_img_crop.width
                    excel_img_crop.height = pil_img_crop.height
                    ws.add_image(excel_img_crop, f"C{ws.max_row}") # C列
                    logger.info(f"裁剪图片 {crop_image_full_path} 已添加到Excel的 C{ws.max_row} 单元格。")
                except Exception as e:
                    logger.error(f"处理裁剪图像 {crop_image_full_path} 失败: {e}", exc_info=True)
            else:
                logger.warning(f"裁剪图片文件未找到: {crop_image_full_path}. 跳过此图片。")

        # 添加完整帧图像
        if person.full_frame_image_path:
            # 移除 person.full_frame_image_path 中重复的 'database/full_frames/' 前缀
            relative_path_from_full_frames_root = person.full_frame_image_path
            if relative_path_from_full_frames_root.startswith("database/full_frames/"):
                relative_path_from_full_frames_root = relative_path_from_full_frames_root[len("database/full_frames/"):]
            elif relative_path_from_full_frames_root.startswith("database\\full_frames\\"):
                relative_path_from_full_frames_root = relative_path_from_full_frames_root[len("database\\full_frames\\"):]

            full_frame_image_full_path = os.path.join(settings.DATABASE_FULL_FRAMES_DIR, relative_path_from_full_frames_root)
            logger.info(f"尝试为人物 {person.uuid} 加载完整帧图片: {full_frame_image_full_path}")

            if os.path.exists(full_frame_image_full_path):
                try:
                    pil_img_full = PILImage.open(full_frame_image_full_path)
                    pil_img_full.thumbnail((settings.EXCEL_EXPORT_IMAGE_SIZE_PX, settings.EXCEL_EXPORT_IMAGE_SIZE_PX))
                    img_byte_arr_full = io.BytesIO()
                    pil_img_full.convert("RGB").save(img_byte_arr_full, format='PNG')
                    img_byte_arr_full.seek(0)
                    excel_img_full = ExcelImage(img_byte_arr_full)
                    excel_img_full.width = pil_img_full.width
                    excel_img_full.height = pil_img_full.height
                    ws.add_image(excel_img_full, f"D{ws.max_row}") # D列
                    logger.info(f"完整帧图片 {full_frame_image_full_path} 已添加到Excel的 D{ws.max_row} 单元格。")
                except Exception as e:
                    logger.error(f"处理完整帧图像 {full_frame_image_full_path} 失败: {e}", exc_info=True)
            else:
                logger.warning(f"完整帧图片文件未找到: {full_frame_image_full_path}. 跳过此图片。")
        
        # 设置行高以容纳图片
        ws.row_dimensions[ws.max_row].height = settings.EXCEL_EXPORT_ROW_HEIGHT_PT

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    logger.info("全部人物特征Excel报告已生成。")
    
    file_name = f"全部人物特征图库_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    generic_filename = "all_persons_features_export_report.xlsx"
    encoded_full_filename = quote(file_name)

    headers = {
        'Content-Disposition': f"attachment; filename=\"{generic_filename}\"; filename*=UTF-8''{encoded_full_filename}\"",
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    }
    return StreamingResponse(io.BytesIO(excel_file.getvalue()), headers=headers, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@router.get("/person_archives", summary="导出人员档案到Excel (仅包含有身份证号/ID且已审核的人物)")
async def export_person_archives(
    db: Session = Depends(get_db),
    # 默认筛选条件，确保只导出有身份证号/ID且已审核的人物
    has_id_card: Optional[bool] = Query(True, description="筛选是否有身份证号/ID数据 (True: 有, False: 无)"),
    is_verified: Optional[bool] = Query(True, description="按审核状态筛选 (True: 已审核, False: 未审核)"),
    marked_for_retrain: Optional[bool] = Query(None, description="按再训练标记筛选 (True: 待再训练, False: 不待再训练)"),
    query: Optional[str] = Query(None, description="根据人物姓名、UUID、身份证号、视频UUID、视频流UUID、图片UUID或名称进行模糊搜索"),
    current_user: schemas.User = Depends(auth.get_current_active_user)
):
    if current_user.role != "admin" and current_user.role != "advanced":
        logger.warning(f"用户 {current_user.username} (角色: {current_user.role}) 未经授权尝试导出人员档案。")
        raise HTTPException(status_code=403, detail="Not enough permissions")

    # 获取所有符合筛选条件的人物数据 (不分页)
    # 强制 is_verified=True 和 has_id_card=True
    persons = crud.get_persons(db, skip=0, limit=None, is_verified=True, marked_for_retrain=marked_for_retrain, query=query, has_id_card=True)

    if not persons:
        logger.info("没有符合筛选条件的人物特征可导出。")
        raise HTTPException(status_code=204, detail="没有符合筛选条件的人物特征可导出")
    
    # 创建新的Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "人员档案"

    # 设置表头
    headers = ["序号", "姓名", "身份证号/ID", "人物UUID", "裁剪图像", "完整帧图像", "来源类型", "来源名称", "来源UUID", "入库时间", "置信度", "是否审核", "是否待再训练"]
    ws.append(headers)

    # 设置列宽
    ws.column_dimensions['B'].width = 20 # 姓名
    ws.column_dimensions['C'].width = 25 # 身份证号/ID
    ws.column_dimensions['D'].width = 38 # 人物UUID
    ws.column_dimensions['E'].width = 15 # 裁剪图像 (原C列)
    ws.column_dimensions['F'].width = 15 # 完整帧图像 (原D列)
    ws.column_dimensions['G'].width = 15 # 来源类型 (原E列)
    ws.column_dimensions['H'].width = 30 # 来源名称 (原F列)
    ws.column_dimensions['I'].width = 38 # 来源UUID (原G列)
    ws.column_dimensions['J'].width = 25 # 入库时间 (原H列)
    ws.column_dimensions['K'].width = 12 # 置信度 (原I列)
    ws.column_dimensions['L'].width = 10 # 是否审核 (原J列)
    ws.column_dimensions['M'].width = 15 # 是否待再训练 (原K列)

    logger.info("Excel 表头和列宽已设置。")

    # 写入数据
    for idx, person in enumerate(persons):
        source_type = "未知"
        source_name = "N/A"
        source_uuid = "N/A"

        if person.stream_uuid:
            source_type = "视频流"
            source_name = person.stream_name or "N/A"
            source_uuid = person.stream_uuid
        elif person.video_uuid:
            source_type = "上传视频"
            source_name = person.video_name or "N/A"
            source_uuid = person.video_uuid
        elif person.upload_image_uuid:
            source_type = "上传图片"
            source_name = person.upload_image_filename or "N/A"
            source_uuid = person.upload_image_uuid
        
        display_time = person.created_at.strftime('%Y-%m-%d %H:%M:%S') if person.created_at else 'N/A'
        confidence_score = f"{person.confidence_score * 100:.2f}%" if person.confidence_score is not None else 'N/A'

        row_data = [
            idx + 1,
            person.name or '',
            person.id_card or '',
            person.uuid,
            '', # 裁剪图像占位符
            '', # 完整帧图像占位符
            source_type,
            source_name,
            source_uuid,
            display_time,
            confidence_score,
            "是" if person.is_verified else "否",
            "是" if person.marked_for_retrain else "否"
        ]
        ws.append(row_data)

        # 添加裁剪图像
        if person.crop_image_path and os.path.exists(os.path.join(settings.BASE_DIR, "backend", person.crop_image_path)):
            try:
                crop_image_full_path = os.path.join(settings.BASE_DIR, "backend", person.crop_image_path)
                pil_img_crop = PILImage.open(crop_image_full_path)
                pil_img_crop.thumbnail((settings.EXCEL_EXPORT_IMAGE_SIZE_PX, settings.EXCEL_EXPORT_IMAGE_SIZE_PX))
                img_byte_arr_crop = io.BytesIO()
                pil_img_crop.convert("RGB").save(img_byte_arr_crop, format='PNG')
                img_byte_arr_crop.seek(0)
                excel_img_crop = ExcelImage(img_byte_arr_crop)
                excel_img_crop.width = pil_img_crop.width
                excel_img_crop.height = pil_img_crop.height
                ws.add_image(excel_img_crop, f"E{ws.max_row}") # E列 (原C列)
                logger.info(f"裁剪图片 {person.crop_image_path} 已添加到Excel的 E{ws.max_row} 单元格。")
            except Exception as e:
                logger.error(f"处理裁剪图像 {person.crop_image_path} 失败: {e}", exc_info=True)
        else:
            logger.warning(f"人物 {person.uuid} 没有裁剪图像路径或文件不存在: {person.crop_image_path}. 跳过添加裁剪图像。")

        # 添加完整帧图像
        if person.full_frame_image_path and os.path.exists(os.path.join(settings.BASE_DIR, "backend", person.full_frame_image_path)):
            try:
                full_frame_image_full_path = os.path.join(settings.BASE_DIR, "backend", person.full_frame_image_path)
                pil_img_full = PILImage.open(full_frame_image_full_path)
                pil_img_full.thumbnail((settings.EXCEL_EXPORT_IMAGE_SIZE_PX, settings.EXCEL_EXPORT_IMAGE_SIZE_PX))
                img_byte_arr_full = io.BytesIO()
                pil_img_full.convert("RGB").save(img_byte_arr_full, format='PNG')
                img_byte_arr_full.seek(0)
                excel_img_full = ExcelImage(img_byte_arr_full)
                excel_img_full.width = pil_img_full.width
                excel_img_full.height = pil_img_full.height
                ws.add_image(excel_img_full, f"F{ws.max_row}") # F列 (原D列)
                logger.info(f"完整帧图片 {person.full_frame_image_path} 已添加到Excel的 F{ws.max_row} 单元格。")
            except Exception as e:
                logger.error(f"处理完整帧图像 {person.full_frame_image_path} 失败: {e}", exc_info=True)
        else:
            logger.warning(f"人物 {person.uuid} 没有完整帧图像路径或文件不存在: {person.full_frame_image_path}. 跳过添加完整帧图像。")
        
        # 设置行高以容纳图片
        ws.row_dimensions[ws.max_row].height = settings.EXCEL_EXPORT_ROW_HEIGHT_PT

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    logger.info("人员档案Excel报告已生成。")
    
    file_name = f"人员档案_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    generic_filename = "person_archives_export_report.xlsx"
    encoded_full_filename = quote(file_name)

    headers = {
        'Content-Disposition': f"attachment; filename=\"{generic_filename}\"; filename*=UTF-8''{encoded_full_filename}\"",
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    }
    return StreamingResponse(io.BytesIO(excel_file.getvalue()), headers=headers, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@router.get("/streams/export_results/{stream_uuid}", summary="导出视频流分析结果到Excel")
async def export_stream_results(
    stream_uuid: str,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: schemas.User = Depends(auth.get_current_active_user)
):
    logger.info(f"DEBUG: Entering export_stream_results for stream_uuid: {stream_uuid}") # 新增这行，确认函数执行

    # 获取流数据
    stream = crud.get_stream_by_uuid(db, stream_uuid)
    if not stream:
        logger.warning(f"未找到视频流: {stream_uuid}")
        raise HTTPException(status_code=404, detail="视频流未找到")

    # 获取与该流相关联的所有人物数据
    persons = crud.get_all_persons_by_stream_id(db, stream_id=stream.id)
    logger.info(f"视频流 (UUID: {stream_uuid}) 找到 {len(persons)} 个人物数据用于导出。")

    # 限制最多导出图片数量，从配置中获取
    persons_to_export = persons[:settings.EXCEL_EXPORT_MAX_IMAGES]
    logger.info(f"实际导出前 {len(persons_to_export)} 张图片。")

    # 创建新的Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "视频流分析结果"

    # 设置表头
    headers = ["序号", "人物UUID", "图片缩小图"]
    ws.append(headers)
    # 设置C列的宽度以容纳图片
    ws.column_dimensions['C'].width = 15 # 假设图片宽度为100像素，约等于15个字符宽度
    logger.info("Excel 表头和列宽已设置。")

    # 写入数据
    for idx, person in enumerate(persons_to_export):
        row_data = [
            idx + 1,
            person.uuid,
        ]
        ws.append(row_data)
        logger.info(f"正在处理人物: {person.uuid}, 裁剪图像路径: {person.crop_image_path}")

        # 添加图片
        # 检查 person.crop_image_path 是否存在
        if not person.crop_image_path:
            logger.warning(f"人物 {person.uuid} 没有 crop_image_path，跳过添加图片。")
            ws.append(["", "", "无缩小图"]) # 添加占位符
            continue

        # 构建图片绝对路径
        logger.info(f"DEBUGGING PATH IN {__file__}")
        logger.info(f"DEBUG: Initial person.crop_image_path: {person.crop_image_path}")
        logger.info(f"DEBUG: settings.DATABASE_CROPS_DIR: {settings.DATABASE_CROPS_DIR}")

        # 统一路径分隔符，将所有反斜杠转换为正斜杠，便于处理
        temp_path = person.crop_image_path.replace("\\", "/")
        logger.info(f"DEBUG: Normalized path (temp_path): {temp_path}")

        # 移除重复的 'database/crops/' 前缀
        prefix = "database/crops/"
        if temp_path.startswith(prefix):
            final_relative_path = temp_path[len(prefix):]
            logger.info(f"DEBUG: After removing prefix ({prefix}). Final relative path: {final_relative_path}")
        else:
            final_relative_path = temp_path # 如果不以预期前缀开头，使用原始归一化路径
            logger.info(f"DEBUG: No expected prefix found, using normalized path as is. Final relative path: {final_relative_path}")

        image_path = os.path.join(settings.DATABASE_CROPS_DIR, final_relative_path)
        logger.info(f"构建的图片绝对路径: {image_path}")
        logger.info(f"图片文件是否存在: {os.path.exists(image_path)}")

        # 检查图片文件是否存在
        if not os.path.exists(image_path):
            logger.warning(f"图片文件未找到: {image_path}. 跳过此图片。")
            ws.append(["", "", "图片文件未找到"]) # 添加占位符
            continue

        # 打开图片并转换为字节流
        try:
            with open(image_path, "rb") as f:
                img_byte_arr = f.read()
            pil_img = PILImage.open(io.BytesIO(img_byte_arr))
            logger.info(f"成功打开图片: {image_path}，大小: {pil_img.size}")
        except Exception as e:
            logger.error(f"处理图片 {image_path} 失败: {e}")
            ws.append(["", "", "图片处理失败"]) # 添加占位符
            continue

        # 缩小图片到合适大小，从配置中获取
        pil_img.thumbnail((settings.EXCEL_EXPORT_IMAGE_SIZE_PX, settings.EXCEL_EXPORT_IMAGE_SIZE_PX))
        logger.info(f"图片 {image_path} 已缩小到 {pil_img.size[0]}x{pil_img.size[1]}。")
        
        # 确保图片为RGB模式
        pil_img = pil_img.convert("RGB")
        
        # 将图片保存到内存中的字节流
        img_byte_arr = io.BytesIO()
        pil_img.save(img_byte_arr, format='PNG')
        img_byte_arr.seek(0) # 重置流的位置到开头

        # 将字节流传递给 ExcelImage
        excel_img = ExcelImage(img_byte_arr)
        excel_img.width = pil_img.width
        excel_img.height = pil_img.height

        # 将图片插入到第三列，即 C 列
        img_cell = f"C{ws.max_row}"
        ws.add_image(excel_img, img_cell)
        # 设置行高以容纳图片，从配置中获取
        ws.row_dimensions[ws.max_row].height = settings.EXCEL_EXPORT_ROW_HEIGHT_PT
        
        # 不再需要删除临时文件，因为没有创建

    # 创建一个临时文件来保存Excel工作簿
    # 不再使用tempfile，直接使用io.BytesIO并返回StreamingResponse
    excel_file_in_memory = io.BytesIO()
    wb.save(excel_file_in_memory)
    excel_file_in_memory.seek(0)
    logger.info(f"视频流 (UUID: {stream_uuid}) 的Excel报告已生成到内存中。")

    # 设置响应头
    file_name = f"视频流_{stream_uuid}_分析结果_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    logger.info(f"生成的视频流导出文件名: {file_name}")

    # 生成通用ASCII回退文件名
    generic_filename = "stream_export_report.xlsx"
    # 对包含中文和时间戳的完整文件名进行URL编码，用于filename*参数
    encoded_full_filename = quote(file_name)

    headers = {
        'Content-Disposition': f"""attachment; filename="{generic_filename}"; filename*=UTF-8''{encoded_full_filename}""",
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    }
    logger.debug(f"导出视频流结果 - 原始文件名: {file_name}, 通用文件名: {generic_filename}, 编码文件名 (for filename*): {encoded_full_filename}, 响应头: {headers}")

    # 返回Excel文件作为响应
    return StreamingResponse(excel_file_in_memory, headers=headers, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet") 